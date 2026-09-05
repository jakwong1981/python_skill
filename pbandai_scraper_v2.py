#!/usr/bin/env python3
"""
P-Bandai Web Scraper v2 - Uses Playwright (real browser) to bypass Cloudflare
Scrapes product information from P-Bandai HK

Updated 2026-09: the scraper queries the category-based search listing
(?_f_categories=04-004, Gunpla) and falls back to the shop page when the
search endpoint is unavailable — its bot protection intermittently returns
"PAGE NOT AVAILABLE", so listings are loaded with retries.

Installation:
  pip3 install playwright
  python3 -m playwright install chromium

Usage:
  python3 pbandai_scraper_v2.py
"""

import json
import csv
import os
import re
import time

# --- Site URLs (category-based search; the search endpoint's bot protection
# intermittently returns "PAGE NOT AVAILABLE") ---
SEARCH_URL = ("https://p-bandai.com/hk/search?_f_categories=04-004&offset=0"
              "&limit=100&sortType=NewArrival&_f_productStatuses=Waiting,On")
SHOP_URL = "https://p-bandai.com/hk/shop/hobbyonlineshop"

# --- Retry settings for the protected search endpoint ---
SEARCH_ATTEMPTS = 1           # top-up tries after the shop page (raise for more)
SEARCH_RETRY_WAIT_S = 10      # seconds to wait between search attempts
USER_AGENT = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
              "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
MAX_PAGES = 10
HEADLESS = True  # set to False to watch the browser run (some bot checks behave differently)

EXTRACT_JS = """() => {
    const products = [];
    const items = document.querySelectorAll('a[href*="/hk/item/"]');
    const seen = new Set();
    items.forEach(a => {
        const url = a.href;
        if (seen.has(url)) return;
        seen.add(url);
        const name = a.querySelector('p')?.innerText?.trim() || '';
        const allParas = a.querySelectorAll('p');
        const price = allParas.length > 1
            ? allParas[1]?.innerText?.trim().replace(/[\u200b\u200c]/g, ' ')
            : 'N/A';
        const statusEl = a.querySelector('li')?.innerText || '';
        if (name) {
            products.push({ name, price, url, status: statusEl });
        }
    });
    return JSON.stringify(products);
}"""


def _new_context(browser):
    return browser.new_context(
        user_agent=USER_AGENT,
        viewport={"width": 1920, "height": 1080},
        locale="zh-HK",
    )


def _wait_for_listing(page, timeout_s=1000, error_grace_s=10):
    """Poll until product cards appear or the site's error page shows up.

    The "PAGE NOT AVAILABLE" page is the bot-protection challenge. It
    sometimes auto-resolves after a few seconds and redirects to the real
    listing, so keep watching for `error_grace_s` before giving up.
    """
    deadline = time.time() + timeout_s
    error_seen_at = None
    while time.time() < deadline:
        if page.locator('a[href*="/hk/item/"]').count() > 0:
            return True
        try:
            if "PAGE NOT AVAILABLE" in page.title():
                if error_seen_at is None:
                    error_seen_at = time.time()
                elif time.time() - error_seen_at >= error_grace_s:
                    return False
        except Exception:
            pass
        page.wait_for_timeout(1000)
    return False


def _extract_products(page, price_tries=12):
    """Extract product cards, retrying briefly while Global-e injects prices."""
    batch = []
    for _ in range(price_tries):
        batch = json.loads(page.evaluate(EXTRACT_JS))
        if not batch or any(p["price"] != "N/A" for p in batch):
            break
        page.wait_for_timeout(1000)
    return batch


NEXT_PAGE_JS = """() => {
    // Smallest offset=/page= value greater than the one in the current URL
    const cur = { offset: 0, page: 0 };
    (location.href.match(/[?&](offset|page)=(\\d+)/g) || []).forEach(x => {
        const mm = x.match(/(offset|page)=(\\d+)/);
        if (mm) cur[mm[1]] = parseInt(mm[2], 10);
    });
    let best = null;
    document.querySelectorAll('a[href]').forEach(a => {
        const mm = a.href.match(/[?&](offset|page)=(\\d+)/);
        if (!mm) return;
        const o = parseInt(mm[2], 10);
        if (o > cur[mm[1]] && (best === null || o < best.value)) {
            best = { param: mm[1], value: o };
        }
    });
    return best;
}"""

TOTAL_JS = """() => {
    // Site's total-result counter, e.g. "42件" / "42 items"
    const nums = [];
    const re = /(\\d+)\\s*(?:件|results?|items?)/gi;
    let m;
    while ((m = re.exec(document.body.innerText)) !== null) {
        nums.push(parseInt(m[1], 10));
    }
    return nums.length ? Math.max(...nums) : null;
}"""


def _extract_stable(page, stable_s=3, max_wait_s=500):
    """Extract cards, waiting until the lazy-loaded card count stops growing.

    The listing renders a first batch of cards and may reveal more later
    (lazy rendering / scroll loading), so re-extract until the card URLs
    are stable for `stable_s` seconds.
    """
    deadline = time.time() + max_wait_s
    last_urls = None
    stable_since = time.time()
    while time.time() < deadline:
        batch = _extract_products(page, price_tries=1)
        urls = [p["url"] for p in batch]
        if urls == last_urls:
            if time.time() - stable_since >= stable_s:
                break
        else:
            last_urls = urls
            stable_since = time.time()
        page.wait_for_timeout(100)
    return _extract_products(page)


def _load_page(browser, url, attempts=5):
    """Load `url` in a fresh context, retrying while the search endpoint is flaky."""
    for attempt in range(1, attempts + 1):
        context = _new_context(browser)
        page = context.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=5000)
            if _wait_for_listing(page):
                return page
            print(f"   ⚠️ Error page (attempt {attempt}/{attempts})")
            try:
                # The challenge may have set a cookie — retry in the SAME
                # context before discarding it
                page.reload(wait_until="domcontentloaded", timeout=300)
                if _wait_for_listing(page):
                    return page
            except Exception:
                pass
        except Exception as e:
            print(f"   ⚠️ Load failed (attempt {attempt}/{attempts}): {type(e).__name__}")
        context.close()
        if attempt < attempts:
            time.sleep(10)
    return None


CJK_RE = re.compile(r"[\u3000-\u9fff]")

# Optional: MyMemory raises its free daily quota from ~5k to 50k chars/day
# when you pass an email address.
MYMEMORY_EMAIL = ""


def _split_delivery_suffix(name):
    """Split a trailing bracketed delivery note (e.g. '[2026年11月發送]') off a name."""
    m = re.search(r"(\s*\[[^\]]*\]\s*)$", name)
    if m:
        return name[:m.start()], m.group(1)
    return name, ""


def _valid_translation(text, original):
    """A translation must differ from the input and contain Chinese.

    Translation services silently echo the input back when their daily
    quota is exhausted; such results must not count as translations.
    """
    if not text:
        return False
    if text.strip().lower() == original.strip().lower():
        return False
    return bool(CJK_RE.search(text))


def _translate_with_backends(text):
    """Translate via Google then MyMemory (3 attempts each); None on failure."""
    try:
        from deep_translator import GoogleTranslator, MyMemoryTranslator
    except ImportError:
        return None
    backends = [
        lambda: GoogleTranslator(source="en", target="zh-TW"),
        lambda: MyMemoryTranslator(
            source="en-GB", target="zh-TW", email=MYMEMORY_EMAIL or None
        ),
    ]
    for make in backends:
        try:
            translator = make()
        except Exception:
            continue
        for _ in range(3):
            try:
                result = translator.translate(text)
                if _valid_translation(result, text):
                    return result
            except Exception:
                pass
            time.sleep(1.5)
    return None


def translate_name(name):
    """Translate an English product name to Traditional Chinese.

    Translation services refuse some trademark strings (e.g. "S.H.Figuarts")
    and echo the input back unchanged. Besides retrying, progressively
    simpler variants are tried — dropping leading dotted brand tokens and
    bracketed qualifiers — and the untouched prefixes are prepended to the
    result. Returns None when everything fails or the library is missing.
    """
    seen = set()
    worklist = [(name, "")]
    while worklist:
        text, prefix = worklist.pop(0)
        if text in seen:
            continue
        seen.add(text)
        result = _translate_with_backends(text)
        if result:
            return prefix + result
        for m in (re.match(r"^(?:(?:[A-Z0-9]+\.)+[A-Za-z0-9]*)\s*(.*)$", text),
                  re.match(r"^\s*(\([^)]*\))\s*(.*)$", text)):
            if m and m.group(1) and m.group(1) != text:
                worklist.append((m.group(1), prefix + text[: m.start(1)]))
    return None


def scrape_pbandai(search_url=SEARCH_URL, shop_url=SHOP_URL):
    """Scrape products using Playwright (real browser = no Cloudflare issues).

    Scrapes the shop page first (statically served, reliable), then tries the
    bot-protected search listing to top up with the full category's items.
    """
    from playwright.sync_api import sync_playwright

    products = []
    seen_urls = set()

    def add_products(batch):
        added = 0
        for prod in batch:
            if prod["url"] in seen_urls:
                continue
            seen_urls.add(prod["url"])
            products.append(prod)
            added += 1
        return added

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)

        # --- Primary: shop page (statically served, not bot-protected) ---
        print(f"🔗 Opening: {shop_url}\n")
        context = _new_context(browser)
        page = context.new_page()
        shop_ok = False
        try:
            page.goto(shop_url, wait_until="domcontentloaded", timeout=30000)
            shop_ok = _wait_for_listing(page)
        except Exception as e:
            print(f"⚠️  Shop page load failed: {type(e).__name__}: {e}\n")
            context.close()
        if shop_ok:
            print("✓ Shop page loaded\n")
            add_products(_extract_products(page))
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(1500)
            extra = add_products(_extract_stable(page, stable_s=2, max_wait_s=6))
            if extra:
                print(f"   Scroll: +{extra} cards\n")
        else:
            print("⚠️  Shop page unavailable\n")

        # --- Top-up: protected search endpoint (may be bot-blocked) ---
        search_ok = False
        for attempt in range(1, SEARCH_ATTEMPTS + 1):
            context = _new_context(browser)
            page = context.new_page()
            try:
                print(f"🔗 Opening: {search_url} "
                      f"(attempt {attempt}/{SEARCH_ATTEMPTS})\n")
                page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
                if _wait_for_listing(page):
                    search_ok = True
                    break
                print("⚠️  Site returned 'PAGE NOT AVAILABLE' (attempt "
                      f"{attempt}/{SEARCH_ATTEMPTS})\n")
                context.close()
            except Exception as e:
                print(f"⚠️  Load failed: {type(e).__name__}: {e} "
                      f"(attempt {attempt}/{SEARCH_ATTEMPTS})\n")
                context.close()
            time.sleep(SEARCH_RETRY_WAIT_S)

        if search_ok:
            print("✓ Search page loaded\n")
            batch = _extract_stable(page)
            print(f"   Page 1: {len(batch)} cards")
            add_products(batch)

            for page_num in range(1, MAX_PAGES):
                # The listing may reveal more cards on scroll — check before paging
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                page.wait_for_timeout(1500)
                batch = _extract_stable(page, stable_s=2, max_wait_s=6)
                added = add_products(batch)
                if added:
                    print(f"   Scroll: +{added} cards (page now {len(batch)})")

                # Stop early when the site's total counter says we have everything
                total = page.evaluate(TOTAL_JS)
                if total is not None and len(products) >= total:
                    print(f"   Captured all {len(products)} of {total} items — done\n")
                    break

                # Next page: follow an explicit link, or construct the URL ourselves
                next_link = page.evaluate(NEXT_PAGE_JS)
                if next_link:
                    key, val = next_link["param"], next_link["value"]
                else:
                    m = re.search(r"[?&](offset|page)=(\d+)", page.url)
                    key = m.group(1) if m else "offset"
                    current = int(m.group(2)) if m else 0
                    val = current + (1 if key == "page" else max(len(batch), 1))
                next_page_url = (re.sub(rf"{key}=\d+", f"{key}={val}", page.url)
                                 if re.search(rf"[?&]{key}=\d+", page.url)
                                 else f"{page.url}&{key}={val}")

                new_page = _load_page(browser, next_page_url)
                if new_page is None:
                    print("   ⚠️ Next page unavailable after retries — stopping\n")
                    break
                page = new_page
                batch = _extract_stable(page)
                added = add_products(batch)
                print(f"   Page {page_num + 1} ({key}={val}): "
                      f"{len(batch)} cards, +{added} new")
                if added == 0:
                    print("   No new items — listing exhausted\n")
                    break
        else:
            print("⚠️  Search endpoint unavailable — keeping shop-page results only\n")

        browser.close()

    print(f"📦 Found {len(products)} products\n")
    return products

def save_json(products, output_dir=None):
    """Save to JSON file (fixed filename, overwritten on each run)"""
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        filename = os.path.join(output_dir, "pbandai_products.json")
    else:
        filename = "pbandai_products.json"
    
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    
    return filename

def save_csv(products, output_dir=None):
    """Save to CSV file (fixed filename, overwritten on each run)"""
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        filename = os.path.join(output_dir, "pbandai_products.csv")
    else:
        filename = "pbandai_products.csv"
    
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=['name', 'tc_name', 'price', 'status', 'url'])
        writer.writeheader()
        writer.writerows(products)
    
    return filename

def save_xlsx(products, output_dir=None):
    """Save to Excel file (fixed filename, overwritten on each run; requires openpyxl)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("⚠️  openpyxl not installed. Run: pip3 install openpyxl")
        return None
    
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        filename = os.path.join(output_dir, "pbandai_products.xlsx")
    else:
        filename = "pbandai_products.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'P-Bandai Products'
    
    # Headers
    headers = ['#', 'Product Name', 'TC Product Name', 'Price (HKD)', 'Status', 'URL']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for i, p in enumerate(products, 1):
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=p['name'])
        ws.cell(row=i+1, column=3, value=p.get('tc_name') or '')
        price_clean = re.sub(r'[^\d.]', '', p['price'])
        ws.cell(row=i+1, column=4, value=float(price_clean) if price_clean.replace('.', '').isdigit() else p['price'])
        ws.cell(row=i+1, column=5, value=p['status'])
        cell = ws.cell(row=i+1, column=6, value=p['url'])
        cell.font = Font(color='0563C1', underline='single')
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 75
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 60
    
    wb.save(filename)
    return filename

def main():
    print("=" * 70)
    print("  P-BANDAI HK PRODUCT SCRAPER v2 (Playwright)")
    print("=" * 70 + "\n")
    
    # Scrape P-Bandai HK products (shop page first, search listing as top-up)
    products = scrape_pbandai()
    
    if products:
        # Translate product names to Traditional Chinese (deep-translator)
        print("\n🌐 Translating product names (en → zh-TW)...")
        failures = 0
        failed_idx = []
        for i, p in enumerate(products, 1):
            core, suffix = _split_delivery_suffix(p["name"])
            if CJK_RE.search(core):
                p["tc_name"] = p["name"]  # already Chinese
            else:
                tc = translate_name(core)
                if tc:
                    p["tc_name"] = tc + suffix
                else:
                    p["tc_name"] = p["name"]
                    failures += 1
                    failed_idx.append(i - 1)
            if i % 10 == 0:
                print(f"   {i}/{len(products)}")
            time.sleep(0.4)

        # Second pass: retry names that failed (transient rate limits)
        for idx in failed_idx:
            core, suffix = _split_delivery_suffix(products[idx]["name"])
            tc = translate_name(core)
            if tc:
                products[idx]["tc_name"] = tc + suffix
                failures -= 1
            time.sleep(3)

        print(f"   Done ({len(products)} names, {failures} untranslated)\n")
        
        # Save files
        json_file = save_json(products)
        csv_file = save_csv(products)
        xlsx_file = save_xlsx(products)
        
        print(f"\n✅ Files saved:")
        print(f"   📄 JSON: {json_file}")
        print(f"   📄 CSV:  {csv_file}")
        if xlsx_file:
            print(f"   📊 XLSX: {xlsx_file}")
        
        # Print summary
        print(f"\n{'='*70}")
        print(f"📊 TOTAL: {len(products)} products")
        print(f"{'='*70}\n")
        
        for i, p in enumerate(products[:5], 1):
            print(f"  {i}. {p['name'][:65]}")
            print(f"     🀄 {p['tc_name'][:50]}")
            print(f"     💰 {p['price']}")
        
        if len(products) > 5:
            print(f"\n  ... and {len(products)-5} more")
    else:
        print("\n❌ No products found.")
        print("   Check the URL or site accessibility.\n")

if __name__ == "__main__":
    main()
