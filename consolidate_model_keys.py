#!/usr/bin/env python3
"""
consolidate_model_keys.py — Scan this Mac for AI-model / API keys, show where
each was found AND which agent/application uses it, consolidated into one
Excel sheet for review.

Usage:
    python3 consolidate_model_keys.py                  # masked values (safe default)
    python3 consolidate_model_keys.py --full           # full secret values in the sheet
    python3 consolidate_model_keys.py --out ~/Desktop/keys.xlsx

Security: by default only MASKED values are written (first4…last4 + length),
enough to identify a key without spreading the secret. Use --full only if
you really need the raw values (the file is then chmod 600).
"""

import argparse
import glob
import json
import os
import re
import shutil
import stat
import sys
from pathlib import Path

HOME = Path.home()

# ── Providers we can recognise from the variable name ────────────────────
PROVIDER_HINTS = {
    "OPENAI": "OpenAI", "AZURE_OPENAI": "Azure OpenAI", "ANTHROPIC": "Anthropic",
    "DEEPSEEK": "DeepSeek", "GEMINI": "Google Gemini", "GOOGLE_AI": "Google AI",
    "GOOGLE_API": "Google", "XAI": "xAI (Grok)", "GROQ": "Groq",
    "MISTRAL": "Mistral", "OPENROUTER": "OpenRouter", "TOGETHER": "Together AI",
    "HUGGINGFACE": "Hugging Face", "HF_": "Hugging Face", "COHERE": "Cohere",
    "PERPLEXITY": "Perplexity", "REPLICATE": "Replicate", "ELEVENLABS": "ElevenLabs",
    "MINIMAX": "MiniMax", "MOONSHOT": "Moonshot", "ZHIPU": "Zhipu AI",
    "DASHSCOPE": "Alibaba DashScope", "QWEN": "Alibaba Qwen", "BAIDU": "Baidu",
    "QIANFAN": "Baidu Qianfan", "SILICONFLOW": "SiliconFlow", "FIREWORKS": "Fireworks AI",
    "PORTKEY": "Portkey", "GITHUB_TOKEN": "GitHub", "GITHUB": "GitHub",
    "GENERATIVE_AI": "Google AI", "VERTEX": "Google Vertex", "LANGCHAIN": "LangChain",
    "LANGSMITH": "LangSmith", "STABILITY": "Stability AI", "RUNPOD": "RunPod",
    "MODAL": "Modal", "SCALE": "Scale AI", "VOYAGE": "Voyage AI", "JINA": "Jina AI",
    "NVIDIA": "NVIDIA", "NEBIUS": "Nebius", "HERMES": "Hermes/Nous",
    "NOUS": "Nous Research", "ARK": "Volcano Engine (ARK)",
    "BYTEPLUS": "Volcano Engine (BytePlus)", "VOLC": "Volcano Engine",
}

# Known agent CLIs + their config dirs + the binary name (if any)
AGENT_DIRS = {
    "Hermes Agent": (".hermes", "hermes"),
    "Claude Code": (".claude", "claude"),
    "Codex (OpenAI)": (".codex", "codex"),
    "OpenCode": (".opencode", "opencode"),
    "Gemini CLI": (".gemini", "gemini"),
    "Cline": (".cline", None),
    "Cursor": (".cursor", None),
    "Windsurf": (".codeium", None),
    "Qoder": (".qoder", None),
    "Antigravity IDE": (".antigravity-ide", None),
    "Continue": (".continue", None),
    "OpenClaw": (".openclaw", "openclaw"),
    "Clawdbot": (".clawdbot", "clawdbot"),
}

# Rough key-name prefix → agents that typically consume that provider
PREFIX_AGENTS = {
    "ANTHROPIC": ["Claude Code", "Hermes Agent"],
    "OPENAI": ["Codex (OpenAI)", "Hermes Agent", "Cline"],
    "GEMINI": ["Gemini CLI", "Hermes Agent", "Cline"],
    "GOOGLE": ["Gemini CLI", "Hermes Agent"],
    "DEEPSEEK": ["Hermes Agent", "OpenCode"],
    "XAI": ["Hermes Agent"],
    "ARK": ["Volcano Engine tools", "OpenCode"],
    "OPENROUTER": ["OpenRouter users", "Cline", "Hermes Agent"],
    "HERMES": ["Hermes Agent"],
}

# Files / globs to scan (read-only)
SCAN_PATHS = [
    ".env",
    ".hermes/.env",
    ".hermes/auth.json",
    ".hermes/**/.env",
    ".hermes/**/auth.json",
    ".hermes/**/credentials*.json",
    ".zshrc", ".zprofile", ".bashrc", ".bash_profile", ".profile",
    "github/**/.env",
    ".config/**/.env",
    ".claude*/**/.env",
    ".codex/**/.env",
    ".openclaw/**/*.env",
    ".openclaw/openclaw.json",
    ".openclaw/credentials/**",
    ".clawdbot/**/.env",
    ".aws/credentials",
    ".netrc",
]

KEY_NAME_RE = re.compile(
    r"^[A-Za-z0-9_]*(?:API[_-]?KEY|TOKEN|SECRET|PASSWORD|PASSWD|CREDENTIAL"
    r"|ACCESS[_-]?KEY|AUTH[_-]?KEY|PRIVATE[_-]?KEY)[A-Za-z0-9_]*$"
)
VALUE_HINT_RE = re.compile(
    r"^(sk-[A-Za-z0-9_\-]+|sk-ant-[A-Za-z0-9_\-]+|AIza[A-Za-z0-9_\-]+"
    r"|gsk_[A-Za-z0-9]+|xai-[A-Za-z0-9_\-]+|ghp_[A-Za-z0-9]+|gho_[A-Za-z0-9]+"
    r"|hf_[A-Za-z0-9]+|r8_[A-Za-z0-9]+|pvc[A-Za-z0-9]+|[A-Za-z0-9_\-]{20,}$)"
)

JSON_KEY_HINT_RE = re.compile(
    r"(api[_-]?key|token|secret|credential|passwd|password|access[_-]?token)", re.I
)


def mask(value, full=False):
    if full:
        return value
    if not value:
        return ""
    if len(value) <= 8:
        return f"({len(value)} chars)"
    return f"{value[:4]}…{value[-4:]} ({len(value)} chars)"


def guess_provider(name):
    upper = name.upper()
    for hint, provider in PROVIDER_HINTS.items():
        if hint in upper:
            return provider
    return "其他 / Other"


# ── Agent detection ───────────────────────────────────────────────────────

def detect_installed_agents():
    """Return names of agents present on this machine (dir or binary)."""
    present = []
    for agent, (dirname, binary) in AGENT_DIRS.items():
        if (HOME / dirname).exists():
            present.append(agent)
        elif binary and shutil.which(binary):
            present.append(agent)
    return present


def load_hermes_providers():
    """Parse ~/.hermes/config.yaml → (default_provider, {key_env: provider_name})."""
    default_provider, key_env_map = None, {}
    cfg_path = HOME / ".hermes" / "config.yaml"
    if not cfg_path.exists():
        return default_provider, key_env_map
    try:
        import yaml  # optional; falls back to regex below
        cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8", errors="replace"))
        model = cfg.get("model") or {}
        default_provider = model.get("provider")
        for cp in cfg.get("custom_providers") or []:
            env_key = cp.get("key_env")
            if env_key:
                key_env_map[env_key] = cp.get("name", env_key)
    except Exception:
        # Regex fallback for key_env=... and provider: ...
        text = cfg_path.read_text(encoding="utf-8", errors="replace")
        m = re.search(r"^\s*provider:\s*(\S+)", text, re.M)
        if m:
            default_provider = m.group(1)
        for m in re.finditer(r"key_env:\s*(\S+)", text):
            key_env_map[m.group(1)] = m.group(1)
    return default_provider, key_env_map


SKIP_DIR_PARTS = {
    "node_modules", ".git", "venv", "lib", "site-packages", "hermes-agent",
    "dist", ".cache", "logs", "__pycache__", ".tmp", "build", "target", ".venv",
}
EVIDENCE_SUFFIXES = (".json", ".toml", ".yaml", ".yml", ".md", ".env", ".txt")


def find_key_references(key_name, skip_path=None):
    """Search installed-agent config dirs for files mentioning this key name.
    Returns a list of "Agent: rel/path" evidence strings (pruned walk)."""
    evidence = []
    for agent, (dirname, _binary) in AGENT_DIRS.items():
        root = HOME / dirname
        if not root.exists():
            continue
        for dirpath, dirnames, filenames in os.walk(root):
            dirnames[:] = [d for d in dirnames if d not in SKIP_DIR_PARTS]
            for fn in filenames:
                p = Path(dirpath) / fn
                if skip_path and str(p) == str(skip_path):
                    continue
                if not fn.endswith(EVIDENCE_SUFFIXES):
                    continue
                try:
                    if p.stat().st_size > 200_000:
                        continue
                    if key_name in p.read_text(encoding="utf-8", errors="ignore")[:50_000]:
                        evidence.append(f"{agent}: {p.relative_to(HOME)}")
                except OSError:
                    continue
            if len(evidence) >= 6:
                return evidence
    return evidence[:6]


def attribute_agent(name, source, installed, default_provider, hermes_env_map):
    """Decide which agent(s) use this key: evidence-based, then fallback."""
    rel = str(source).replace(str(HOME), "~")
    upper = name.upper()
    notes = []

    # 0) Config-dir ownership: the file lives inside an agent's config dir
    for agent, (dirname, _binary) in AGENT_DIRS.items():
        if dirname == ".hermes":
            continue
        if str(HOME / dirname) in str(source):
            notes.append(f"{agent} (設定檔: {rel.split('/')[-1]})")

    # 1) Hermes-managed keys
    if "/.hermes/" in rel or rel == "~/.env":
        if name in hermes_env_map:
            notes.append(f"Hermes Agent (custom provider: {hermes_env_map[name]})")
        elif default_provider and default_provider.upper() in upper:
            notes.append(f"Hermes Agent (default provider: {default_provider})")
        else:
            notes.append("Hermes Agent")
        if "auth.json" in rel:
            notes.append("Hermes OAuth/credential store")

    # 2) Shell-profile exports → available to everything the user runs
    if rel.startswith("~/.zshrc") or rel.startswith("~/.bashrc") or \
       rel.startswith("~/.zprofile") or rel.startswith("~/.bash_profile") or rel.startswith("~/.profile"):
        candidates = PREFIX_AGENTS.get(
            next((k for k in PREFIX_AGENTS if upper.startswith(k)), ""), [])
        alive = [a for a in candidates if a in installed] or installed
        notes.append("Shell profile export (可用於所有 terminal agents)")
        notes.append("已安裝候選: " + ", ".join(alive[:6]))

    # 3) Config-file references (strongest evidence)
    for ev in find_key_references(name, skip_path=source):
        notes.append(f"配置引用: {ev}")

    # 4) Fallback by provider prefix
    if not notes:
        candidates = PREFIX_AGENTS.get(
            next((k for k in PREFIX_AGENTS if upper.startswith(k)), ""), [])
        alive = [a for a in candidates if a in installed]
        notes.append(("可能使用: " + ", ".join(alive)) if alive else "未能判定 (請自行確認)")

    return " | ".join(notes[:4])


# ── Scanners ──────────────────────────────────────────────────────────────

def _clean_value(raw):
    """Strip quotes + inline comments from an env value."""
    value = raw.strip()
    if value.startswith('"') or value.startswith("'"):
        q = value[0]
        end = value.find(q, 1)
        value = value[1:end] if end > 0 else value[1:]
    else:
        value = re.split(r"\s+#", value, maxsplit=1)[0]
    return value.strip()


def scan_line_file(path, full):
    """Parse KEY=VALUE style files (.env, rc files)."""
    found = []
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as fh:
            for lineno, line in enumerate(fh, 1):
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                m = re.match(r"^(?:export\s+)?([A-Za-z_][A-Za-z0-9_]*)=(.*)$", line)
                if not m:
                    continue
                name = m.group(1)
                if name.startswith("CONDA_") or name.startswith("NVM_") or name == "PATH":
                    continue
                value = _clean_value(m.group(2))
                if not KEY_NAME_RE.match(name) or value.startswith("#"):
                    continue
                note = ""
                if not value:
                    note = "⚠️ 空值 / 未設定有效 key"
                elif len(value) < 16 and not VALUE_HINT_RE.match(value):
                    note = f"⚠️ 疑似佔位/測試值 ({len(value)} chars)"
                found.append({
                    "name": name, "value": value,
                    "source": str(path), "line": lineno,
                    "provider": guess_provider(name), "note": note,
                })
    except (OSError, UnicodeDecodeError):
        pass
    return found


def walk_json(obj, path_prefix=""):
    """Yield (keypath, value) for string values in nested JSON."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            yield from walk_json(v, f"{path_prefix}.{k}" if path_prefix else str(k))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            yield from walk_json(v, f"{path_prefix}[{i}]")
    elif isinstance(obj, str) and obj:
        yield path_prefix, obj


def scan_json_file(path, full):
    found = []
    try:
        data = json.load(open(path, "r", encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return found
    for keypath, value in walk_json(data):
        leaf = keypath.split(".")[-1].split("[")[0]
        if (JSON_KEY_HINT_RE.search(keypath) and VALUE_HINT_RE.match(value)) or KEY_NAME_RE.match(leaf):
            found.append({
                "name": keypath, "value": value,
                "source": str(path), "line": 0,
                "provider": guess_provider(keypath), "note": "",
            })
    return found


def collect(full):
    results = []
    seen_files = set()
    for pattern in SCAN_PATHS:
        for p in sorted(glob.glob(str(HOME / pattern), recursive=True)):
            path = Path(p)
            if not path.is_file() or str(path) in seen_files:
                continue
            seen_files.add(str(path))
            if p.endswith(".json") or p.endswith("auth.json") or "credentials" in p:
                results.extend(scan_json_file(path, full))
            else:
                results.extend(scan_line_file(path, full))
    # Dedupe: same key name in the same file → keep the last occurrence
    seen = {}
    for r in results:
        seen[(r["source"], r["name"])] = r
    return list(seen.values())


# ── Excel output ──────────────────────────────────────────────────────────

def write_excel(rows, out_path, full):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Model Keys Review"

    headers = ["#", "Key 名稱 / 用途", "Provider / 服務商", "使用者 Agent",
               "來源檔案", "行", "值 (Value)", "備註"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    note = "值已遮罩 (只顯示頭尾4位) — 如需完整值: 加 --full 重新執行" if not full else "⚠️ 完整機密值 — 請妥善保管此檔案"
    for idx, r in enumerate(rows, 1):
        ws.append([idx, r["name"], r["provider"], r["agent"],
                   r["source"], r["line"] or "", mask(r["value"], full),
                   r.get("note") or (note if idx == 1 else "")])
    for col, width in zip(range(1, 9), [5, 38, 20, 52, 46, 6, 40, 34]):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows) + 1}"

    wb.save(out_path)
    try:
        os.chmod(out_path, stat.S_IRUSR | stat.S_IWUSR)  # 600
    except OSError:
        pass


def main():
    ap = argparse.ArgumentParser(description="Consolidate model/API keys + who uses them into one Excel sheet")
    ap.add_argument("--full", action="store_true", help="write full secret values (default: masked)")
    ap.add_argument("--out", default=str(HOME / "Desktop" / "model_keys_review.xlsx"))
    args = ap.parse_args()

    rows = collect(args.full)
    if not rows:
        print("No keys found.")
        return 1

    installed = detect_installed_agents()
    default_provider, hermes_env_map = load_hermes_providers()
    for r in rows:
        r["agent"] = attribute_agent(r["name"], r["source"], installed,
                                     default_provider, hermes_env_map)

    out = os.path.expanduser(args.out)
    write_excel(rows, out, args.full)

    print(f"✅ Found {len(rows)} key(s) → {out}")
    print(f"   遮罩模式: {'OFF (full values)' if args.full else 'ON (masked)'}")
    print(f"   已偵測 agents: {', '.join(installed) if installed else '(none found)'}")
    print("-" * 100)
    for i, r in enumerate(rows, 1):
        src = os.path.basename(r["source"]) or r["source"]
        note = f"  [{r.get('note')}]" if r.get("note") else ""
        print(f"{i:>3}  {r['provider']:<20} {r['name']:<36} {mask(r['value'], args.full):<24} {src}{note}")
        print(f"      ➜ {r['agent']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
